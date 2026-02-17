#!/usr/bin/env python3
"""
Newsletter - Envia Excel com últimas notícias por email
"""
import os
import io
import resend
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_connection


# Configuração
RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
RECIPIENTS = os.environ.get("RECIPIENTS", "").split(",")
FROM_EMAIL = os.environ.get("FROM_EMAIL", "noticias@resend.dev")


def get_recent_news(hours=24, limit=100):
    """Busca notícias das últimas X horas"""
    conn = get_connection()
    cur = conn.cursor()

    cutoff = datetime.utcnow() - timedelta(hours=hours)

    cur.execute("""
        SELECT
            n.title,
            n.link,
            n.category,
            n.priority_score,
            s.name as source_name,
            n.published_at,
            n.fetched_at
        FROM news n
        LEFT JOIN sources s ON n.source_id = s.id
        WHERE n.fetched_at > %s
        ORDER BY n.priority_score DESC, n.fetched_at DESC
        LIMIT %s
    """, (cutoff, limit))

    columns = ['title', 'link', 'category', 'priority_score', 'source', 'published_at', 'fetched_at']
    news = [dict(zip(columns, row)) for row in cur.fetchall()]

    cur.close()
    conn.close()

    return news


def get_stats():
    """Busca estatísticas do banco"""
    conn = get_connection()
    cur = conn.cursor()

    stats = {}

    # Total últimas 24h
    cutoff = datetime.utcnow() - timedelta(hours=24)
    cur.execute("SELECT COUNT(*) FROM news WHERE fetched_at > %s", (cutoff,))
    stats['last_24h'] = cur.fetchone()[0]

    # Por categoria (últimas 24h)
    cur.execute("""
        SELECT category, COUNT(*) as count
        FROM news
        WHERE fetched_at > %s
        GROUP BY category
        ORDER BY count DESC
    """, (cutoff,))
    stats['by_category'] = {row[0]: row[1] for row in cur.fetchall()}

    # Alta prioridade (últimas 24h)
    cur.execute("""
        SELECT COUNT(*) FROM news
        WHERE fetched_at > %s AND priority_score >= 2
    """, (cutoff,))
    stats['high_priority'] = cur.fetchone()[0]

    cur.close()
    conn.close()

    return stats


def create_excel(news, stats):
    """Cria arquivo Excel com as notícias"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Notícias"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalhos
    headers = ['Título', 'Categoria', 'Fonte', 'Score', 'Link']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Dados
    for row, item in enumerate(news, 2):
        ws.cell(row=row, column=1, value=item['title']).border = border
        ws.cell(row=row, column=2, value=item['category']).border = border
        ws.cell(row=row, column=3, value=item['source']).border = border
        ws.cell(row=row, column=4, value=item['priority_score']).border = border
        ws.cell(row=row, column=5, value=item['link']).border = border

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 80  # Título
    ws.column_dimensions['B'].width = 15  # Categoria
    ws.column_dimensions['C'].width = 25  # Fonte
    ws.column_dimensions['D'].width = 8   # Score
    ws.column_dimensions['E'].width = 50  # Link

    # Sheet de estatísticas
    ws_stats = wb.create_sheet("Estatísticas")
    ws_stats.cell(row=1, column=1, value="Métrica").font = header_font
    ws_stats.cell(row=1, column=1).fill = header_fill
    ws_stats.cell(row=1, column=2, value="Valor").font = header_font
    ws_stats.cell(row=1, column=2).fill = header_fill

    ws_stats.cell(row=2, column=1, value="Total (24h)")
    ws_stats.cell(row=2, column=2, value=stats['last_24h'])
    ws_stats.cell(row=3, column=1, value="Alta Prioridade")
    ws_stats.cell(row=3, column=2, value=stats['high_priority'])

    row = 5
    ws_stats.cell(row=row, column=1, value="Por Categoria").font = Font(bold=True)
    row += 1
    for cat, count in stats['by_category'].items():
        ws_stats.cell(row=row, column=1, value=cat)
        ws_stats.cell(row=row, column=2, value=count)
        row += 1

    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 10

    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def create_html_summary(news, stats):
    """Cria resumo HTML para o corpo do email"""
    today = datetime.now().strftime("%d/%m/%Y")

    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; color: #333; }}
            h1 {{ color: #2F5496; }}
            .stats {{ background: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0; }}
            .stat-item {{ display: inline-block; margin-right: 30px; }}
            .stat-value {{ font-size: 24px; font-weight: bold; color: #2F5496; }}
            .stat-label {{ font-size: 12px; color: #666; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th {{ background: #2F5496; color: white; padding: 10px; text-align: left; }}
            td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
            tr:hover {{ background: #f5f5f5; }}
            .high {{ color: #c00; font-weight: bold; }}
            a {{ color: #2F5496; }}
        </style>
    </head>
    <body>
        <h1>Newsletter Politics - {today}</h1>

        <div class="stats">
            <div class="stat-item">
                <div class="stat-value">{stats['last_24h']}</div>
                <div class="stat-label">Notícias (24h)</div>
            </div>
            <div class="stat-item">
                <div class="stat-value">{stats['high_priority']}</div>
                <div class="stat-label">Alta Prioridade</div>
            </div>
        </div>

        <h2>Top 20 Notícias (por prioridade)</h2>
        <table>
            <tr>
                <th>Título</th>
                <th>Categoria</th>
                <th>Fonte</th>
                <th>Score</th>
            </tr>
    """

    for item in news[:20]:
        score_class = 'high' if item['priority_score'] >= 3 else ''
        html += f"""
            <tr>
                <td><a href="{item['link']}">{item['title'][:80]}{'...' if len(item['title']) > 80 else ''}</a></td>
                <td>{item['category']}</td>
                <td>{item['source']}</td>
                <td class="{score_class}">{item['priority_score']:.1f}</td>
            </tr>
        """

    html += """
        </table>
        <p style="margin-top: 20px; color: #666; font-size: 12px;">
            Veja todas as notícias no arquivo Excel em anexo.
        </p>
    </body>
    </html>
    """

    return html


def send_newsletter(hours=24, limit=100):
    """Gera e envia a newsletter"""
    if not RESEND_API_KEY:
        print("ERRO: RESEND_API_KEY não configurada")
        return False

    if not RECIPIENTS or RECIPIENTS == ['']:
        print("ERRO: RECIPIENTS não configurado")
        return False

    # Configurar Resend
    resend.api_key = RESEND_API_KEY

    # Buscar dados
    print(f"Buscando notícias das últimas {hours}h...")
    news = get_recent_news(hours=hours, limit=limit)
    stats = get_stats()

    if not news:
        print("Nenhuma notícia encontrada")
        return False

    print(f"Encontradas {len(news)} notícias")

    # Criar Excel
    print("Gerando Excel...")
    excel_data = create_excel(news, stats)

    # Criar HTML
    html_content = create_html_summary(news, stats)

    # Nome do arquivo
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"noticias_{today}.xlsx"

    # Enviar email
    print(f"Enviando para: {', '.join(RECIPIENTS)}")

    try:
        import base64
        excel_base64 = base64.b64encode(excel_data).decode('utf-8')

        response = resend.Emails.send({
            "from": FROM_EMAIL,
            "to": RECIPIENTS,
            "subject": f"Newsletter Politics - {today}",
            "html": html_content,
            "attachments": [
                {
                    "filename": filename,
                    "content": excel_base64,
                }
            ]
        })

        print(f"Email enviado! ID: {response['id']}")
        return True

    except Exception as e:
        print(f"Erro ao enviar email: {e}")
        return False


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Newsletter Politics")
    parser.add_argument("--hours", "-H", type=int, default=24, help="Horas para buscar (default: 24)")
    parser.add_argument("--limit", "-l", type=int, default=100, help="Limite de notícias (default: 100)")
    parser.add_argument("--test", "-t", action="store_true", help="Modo teste (não envia email)")

    args = parser.parse_args()

    if args.test:
        print("=== MODO TESTE ===")
        news = get_recent_news(hours=args.hours, limit=args.limit)
        stats = get_stats()
        print(f"Notícias: {len(news)}")
        print(f"Stats: {stats}")
        print(f"Recipients: {RECIPIENTS}")

        # Gerar Excel local para teste
        excel_data = create_excel(news, stats)
        with open("test_newsletter.xlsx", "wb") as f:
            f.write(excel_data)
        print("Excel salvo: test_newsletter.xlsx")
    else:
        send_newsletter(hours=args.hours, limit=args.limit)


if __name__ == "__main__":
    main()
